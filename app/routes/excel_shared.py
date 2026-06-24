# app/routes/excel_shared.py
"""Shared Excel helpers for monthly exports.

`populate_month_sheet` renders one agent's month into a worksheet, exactly the layout
used by the per-person month export (/month/{id}/export-excel). It is reused by the
admin monthly report so each agent tab looks identical to the single-person export.
"""

import calendar as _cal
from datetime import date as _date

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Column layout for an agent's month sheet (index used to decide merge and filter).
COL_HEADERS = [
    "Datum",
    "Veckodag",
    "Skifttyp",
    "Start",
    "Slut",
    "Vanliga timmar",
    "OB Vardagskväll(x1,12)",
    "OB Natt(x1,18)",
    "OB tillägg helg(x1,24)",  # OB3 + OB4 merged
    "OB tillägg storhelg(x1,47)",
    "B.Vardag(75)",
    "Beredskap Helg(97)",
    "Beredskap Helgdag(112)",
    "Beredskap Storhelg(192)",
    "Övertid(x2)",
    "Kommentar",
]
# First numeric column (0-based, from COL_HEADERS): "Vanliga timmar".
NUM_START = 5

# Alternative header labels used by the admin monthly report's agent tabs. Same columns
# and order as COL_HEADERS, but OB uses the payroll codes (150-153) like the summary sheet.
REPORT_COL_HEADERS = [
    "Datum",
    "Veckodag",
    "Skifttyp",
    "Start",
    "Slut",
    "Vanliga timmar",
    "Kväll 150",
    "Natt 151",
    "Helg 152",
    "Storhelg 153",
    "Beredskap Vardag 75",
    "Beredskap Helg 97",
    "Beredskap Helgdag 112",
    "Beredskap Storhelg 192",
    "Övertid",
    "Kommentar",
]

_SV_DAYS = ["Måndag", "Tisdag", "Onsdag", "Torsdag", "Fredag", "Lördag", "Söndag"]


def autofit_columns(ws, min_width: int = 4, max_width: int = 40, padding: int = 0) -> None:
    """Size each column to its widest cell, header included, so headers are never clipped.

    Width is the longest value in the column (header or data), clamped to [min, max].
    No extra padding is added: Excel's built-in cell margin already prevents clipping, so
    adding characters here only leaves wasted whitespace on the right.
    """
    for col_cells in ws.columns:
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        width = min(max(longest + padding, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _fmt_time(t) -> str:
    if t is None:
        return ""
    if hasattr(t, "strftime"):
        return t.strftime("%H:%M")
    return str(t)[:5]


def _build_month_rows(
    summary: dict, year: int, month: int, headers: list[str] = COL_HEADERS, split_oncall_overtime: bool = True
):
    """Build (final_headers, final_rows, final_totals) for one agent's month.

    Mirrors the per-person export: one row per calendar day and numeric columns with no
    data dropped (text columns always kept). `headers` selects the displayed labels (same
    structure as COL_HEADERS). With split_oncall_overtime=True a day with both on-call and
    overtime is shown as two rows; with False they share one "Beredskap/Övertid" row.
    """
    num_days_in_month = _cal.monthrange(year, month)[1]
    all_dates = [_date(year, month, d) for d in range(1, num_days_in_month + 1)]
    day_lookup = {d["date"]: d for d in summary.get("days", [])}

    rows = []
    totals = [0.0] * len(COL_HEADERS)

    for day_date in all_dates:
        d = day_lookup.get(day_date)
        shift = d.get("shift") if d else None
        if d and d.get("parental_leave"):
            # Week-based parental leave renders as a LEAVE shift; label it as parental.
            shift_label = "Föräldraledigt"
        elif not shift or shift.code == "OFF":
            shift_label = "Ledig"
        else:
            shift_label = shift.label if hasattr(shift, "label") and shift.label else shift.code

        partial_absence = d.get("partial_absence") if d else None
        comment_parts = []
        if partial_absence:
            if partial_absence.arrived_at:
                comment_parts.append(f"Sen ankomst {partial_absence.arrived_at}")
            if partial_absence.left_at:
                comment_parts.append(f"Slutade tidigt {partial_absence.left_at}")
        comment = ", ".join(comment_parts)

        if d and shift and shift.code not in ("OFF",):
            start_t = d.get("start")
            end_t = d.get("end")
            if not start_t and shift and shift.start_time:
                start_t = shift.start_time
                end_t = shift.end_time
            if d.get("ot_details") and d["ot_details"].get("is_extension"):
                end_t_str = d["ot_details"]["end_time"][:5]
            else:
                end_t_str = _fmt_time(end_t)
            start_str = _fmt_time(start_t)

            ob1 = d["ob_hours"].get("OB1", 0) or 0
            ob2 = d["ob_hours"].get("OB2", 0) or 0
            ob3 = (d["ob_hours"].get("OB3", 0) or 0) + (d["ob_hours"].get("OB4", 0) or 0)
            ob5 = d["ob_hours"].get("OB5", 0) or 0

            ob_sum = ob1 + ob2 + ob3 + ob5
            is_work = shift.code not in ("OC", "OT")
            hours = d.get("hours", 0) or 0
            norm = max((hours - ob_sum), 0) if is_work else 0

            oc_bd = d.get("oncall_details", {}).get("breakdown", {}) if d.get("oncall_details") else {}
            oc_vardag = oc_bd.get("OC_WEEKDAY", {}).get("hours", 0) or 0
            oc_helg = (
                (oc_bd.get("OC_WEEKEND", {}).get("hours", 0) or 0)
                + (oc_bd.get("OC_WEEKEND_SAT", {}).get("hours", 0) or 0)
                + (oc_bd.get("OC_WEEKEND_SUN", {}).get("hours", 0) or 0)
                + (oc_bd.get("OC_WEEKEND_MON", {}).get("hours", 0) or 0)
            )
            oc_helgdag = (
                (oc_bd.get("OC_HOLIDAY", {}).get("hours", 0) or 0)
                + (oc_bd.get("OC_HOLIDAY_EVE", {}).get("hours", 0) or 0)
                + (oc_bd.get("OC_NATIONALDAGEN", {}).get("hours", 0) or 0)
            )
            oc_storhelg = oc_bd.get("OC_SPECIAL", {}).get("hours", 0) or 0
            ot = d.get("ot_hours", 0) or 0

            num_vals = [norm, ob1, ob2, ob3, ob5, oc_vardag, oc_helg, oc_helgdag, oc_storhelg, ot, comment]

            oc_total = oc_vardag + oc_helg + oc_helgdag + oc_storhelg
            if oc_total > 0 and ot > 0 and split_oncall_overtime:
                # Split into an on-call row and an overtime row
                oc_vals = [0, 0, 0, 0, 0, oc_vardag, oc_helg, oc_helgdag, oc_storhelg, 0, ""]
                ot_vals = [norm, ob1, ob2, ob3, ob5, 0, 0, 0, 0, ot, comment]
                rows.append([str(day_date), _SV_DAYS[day_date.weekday()], "Beredskap", "", ""] + oc_vals)
                rows.append([str(day_date), _SV_DAYS[day_date.weekday()], shift_label, start_str, end_t_str] + ot_vals)
            elif oc_total > 0 and ot > 0:
                # Keep on-call and overtime on a single row
                rows.append(
                    [str(day_date), _SV_DAYS[day_date.weekday()], "Beredskap/Övertid", start_str, end_t_str] + num_vals
                )
            else:
                rows.append([str(day_date), _SV_DAYS[day_date.weekday()], shift_label, start_str, end_t_str] + num_vals)
        else:
            num_vals = [0.0] * 10 + [comment]
            rows.append([str(day_date), _SV_DAYS[day_date.weekday()], shift_label, "", ""] + num_vals)

        for i, v in enumerate(num_vals[:-1]):  # exclude comment column
            totals[NUM_START + i] = totals[NUM_START + i] + v

    # Drop numeric columns with no data anywhere; always keep the fixed text columns.
    active_num = [any(row[NUM_START + i] for row in rows) for i in range(len(COL_HEADERS) - NUM_START)]
    keep_cols = list(range(NUM_START)) + [NUM_START + i for i, v in enumerate(active_num) if v]

    final_headers = [headers[i] for i in keep_cols]
    final_rows = [[row[i] for i in keep_cols] for row in rows]
    final_totals = [totals[i] for i in keep_cols]
    return final_headers, final_rows, final_totals


def populate_month_sheet(
    ws,
    summary: dict,
    year: int,
    month: int,
    headers: list[str] = COL_HEADERS,
    split_oncall_overtime: bool = True,
) -> None:
    """Fill a worksheet with one agent's month, matching the per-person Excel export.

    `headers` overrides the displayed column labels (e.g. REPORT_COL_HEADERS for the
    monthly report's agent tabs); the column structure is unchanged. With
    split_oncall_overtime=False, on-call and overtime share one row instead of two.
    """
    final_headers, final_rows, final_totals = _build_month_rows(summary, year, month, headers, split_oncall_overtime)

    header_fill = PatternFill("solid", fgColor="2D3748")
    header_font = Font(bold=True, color="FFFFFF")
    total_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")

    ws.append(final_headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row in final_rows:
        ws.append(row)
        r = ws.max_row
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = "0.0"
                cell.alignment = right_align
                if val == 0:
                    cell.value = None

    # Total row
    total_row = []
    for i, _h in enumerate(final_headers):
        if i == 0:
            total_row.append("Total")
        elif i < NUM_START:
            total_row.append(None)
        else:
            v = final_totals[i]
            total_row.append(round(v, 1) if v else None)

    ws.append(total_row)
    r = ws.max_row
    for cell in ws[r]:
        cell.font = total_font
        if isinstance(cell.value, float):
            cell.number_format = "0.0"
            cell.alignment = right_align

    autofit_columns(ws)
