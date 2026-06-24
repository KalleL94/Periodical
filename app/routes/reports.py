# app/routes/reports.py
"""Admin monthly report: one row per agent and substitute with hours, overtime,
on-call and absence figures, viewable in the browser and exportable as CSV."""

import csv
import io

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.auth.auth import get_admin_user
from app.core.helpers import render_template
from app.core.schedule import build_month_report, rotation_start_date
from app.core.utils import get_safe_today
from app.core.validators import validate_date_params
from app.database.database import User, get_db
from app.routes.shared import templates

router = APIRouter(tags=["reports"])

MONTH_NAMES_SV = [
    "Januari",
    "Februari",
    "Mars",
    "April",
    "Maj",
    "Juni",
    "Juli",
    "Augusti",
    "September",
    "Oktober",
    "November",
    "December",
]

# CSV column order: (dict key, header label)
CSV_COLUMNS = [
    ("person_name", "Namn"),
    ("num_shifts", "Antal pass"),
    ("total_hours", "Timmar"),
    ("ot_hours", "Övertid (h)"),
    ("oncall_hours", "Beredskap (h)"),
    ("ob_kvall", "Kväll 150 (h)"),
    ("ob_natt", "Natt 151 (h)"),
    ("ob_helg", "Helg 152 (h)"),
    ("ob_storhelg", "Storhelg 153 (h)"),
    ("sick_days", "Sjuk (dgr)"),
    ("vab_days", "VAB (dgr)"),
    ("leave_days", "Ledigt (dgr)"),
    ("off_days", "Frånvaro övrigt (dgr)"),
    ("parental_days", "Föräldraledigt (dgr)"),
    ("vacation_days", "Semester (dgr)"),
]


# Excel sheet titles cannot contain []:*?/\ and are limited to 31 characters.
_SHEET_TRANS = str.maketrans({c: " " for c in "[]:*?/\\"})


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Return a valid, unique Excel sheet title derived from a person's name."""
    base = (name or "").translate(_SHEET_TRANS).strip()[:31] or "Agent"
    title = base
    i = 2
    while title in used:
        suffix = f" ({i})"
        title = f"{base[: 31 - len(suffix)]}{suffix}"
        i += 1
    return title


def _resolve_year_month(year: int | None, month: int | None) -> tuple[int, int]:
    safe_today = get_safe_today(rotation_start_date)
    year = year or safe_today.year
    month = month or safe_today.month
    validate_date_params(year, month, None)
    return year, month


@router.get("/admin/report", response_class=HTMLResponse, name="admin_report")
async def admin_report(
    request: Request,
    year: int = None,
    month: int = None,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """Monthly report table for all agents and substitutes."""
    year, month = _resolve_year_month(year, month)
    rows = build_month_report(year, month, db)

    prev_month = month - 1 or 12
    prev_year = year - 1 if month == 1 else year
    next_month = month + 1 if month < 12 else 1
    next_year = year + 1 if month == 12 else year

    return render_template(
        templates,
        "admin_report.html",
        request,
        {
            "year": year,
            "month": month,
            "month_name": MONTH_NAMES_SV[month - 1],
            "rows": rows,
            "prev_year": prev_year,
            "prev_month": prev_month,
            "next_year": next_year,
            "next_month": next_month,
        },
        user=current_user,
    )


@router.get("/admin/report.csv", name="admin_report_csv")
async def admin_report_csv(
    year: int = None,
    month: int = None,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """Download the monthly report as a CSV file (UTF-8 with BOM for Excel)."""
    year, month = _resolve_year_month(year, month)
    rows = build_month_report(year, month, db)

    buffer = io.StringIO()
    buffer.write("﻿")  # BOM so Excel reads UTF-8 (åäö) correctly
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow([label for _, label in CSV_COLUMNS])
    for row in rows:
        writer.writerow([row.get(key, "") for key, _ in CSV_COLUMNS])

    buffer.seek(0)
    filename = f"rapport-{year}-{month:02d}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/admin/report.xlsx", name="admin_report_xlsx")
async def admin_report_xlsx(
    year: int = None,
    month: int = None,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """Download the monthly report as an Excel workbook.

    The first sheet is the consolidated report (same columns as the CSV). Each agent
    (rotation positions 1-10) then gets its own sheet, formatted exactly like the
    per-person month export.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    from app.core.schedule import build_substitute_month_summaries
    from app.core.schedule.summary import summarize_month_for_person
    from app.routes.excel_shared import REPORT_COL_HEADERS, autofit_columns, populate_month_sheet

    year, month = _resolve_year_month(year, month)
    rows = build_month_report(year, month, db)

    wb = openpyxl.Workbook()

    # Sheet 1: consolidated summary, same layout as the CSV.
    ws = wb.active
    ws.title = "Sammanställning"
    ws.append([label for _, label in CSV_COLUMNS])
    header_fill = PatternFill("solid", fgColor="2D3748")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(key, "") for key, _ in CSV_COLUMNS])
    autofit_columns(ws)

    # One sheet per agent (rotation positions 1-10), identical to the /month/{id} export.
    used_titles = {ws.title}
    for pid in range(1, 11):
        summary = summarize_month_for_person(year, month, pid, session=db, payment_year=year)
        title = _safe_sheet_title(summary.get("person_name") or f"Agent {pid}", used_titles)
        used_titles.add(title)
        agent_ws = wb.create_sheet(title=title)
        populate_month_sheet(agent_ws, summary, year, month, headers=REPORT_COL_HEADERS, split_oncall_overtime=False)

    # One sheet per substitute with activity in the month, same layout as the agent tabs.
    for sub_summary in build_substitute_month_summaries(year, month, db, include_overtime=True):
        title = _safe_sheet_title(sub_summary.get("person_name") or "Vikarie", used_titles)
        used_titles.add(title)
        sub_ws = wb.create_sheet(title=title)
        populate_month_sheet(sub_ws, sub_summary, year, month, headers=REPORT_COL_HEADERS, split_oncall_overtime=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"rapport-{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
