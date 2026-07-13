"""Integration tests for the /month team view under mid-month person changes.

The month matrix is publicly viewable (no login needed). Schedule internals read
the rotation era through the global SessionLocal, while routes read PersonHistory
through the get_db override. To make both see the same data, we bind a
monkeypatched SessionLocal to the same in-memory engine as test_db and seed a
RotationEra plus the PersonHistory rows there.
"""

import datetime
import re

import pytest
from sqlalchemy.orm import sessionmaker

import app.database.database as db_module
from app.auth.auth import create_access_token
from app.core.schedule import (
    build_week_data,
    clear_schedule_cache,
    generate_month_data,
    summarize_month_for_person,
    summarize_year_for_person,
)
from app.core.schedule.period import mask_days_to_employment
from app.core.schedule.person_history import add_person_change, end_employment, start_employment, swap_positions
from app.core.utils import get_today
from app.database.database import Absence, AbsenceType, OvertimeShift, RotationEra, User, UserRole, WageType
from tests.conftest import _ROTATION_ERA_PATTERN


def _make_user(session, uid, username, name, *, person_id=None, role=UserRole.USER):
    user = User(
        id=uid,
        username=username,
        password_hash="x",
        name=name,
        role=role,
        wage=30000,
        wage_type=WageType.MONTHLY,
        vacation={},
        must_change_password=0,
        is_active=0,
        person_id=person_id,
    )
    session.add(user)
    session.commit()
    return user


@pytest.fixture()
def month_env(test_db, test_client, monkeypatch):
    """Bind the global SessionLocal to test_db's engine and seed a rotation era.

    Yields (test_client, test_db) so schedule internals and the HTTP route share
    one in-memory database and both resolve the rotation and PersonHistory rows.
    """
    engine = test_db.get_bind()
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    monkeypatch.setattr(db_module, "SessionLocal", SessionLocal)
    clear_schedule_cache()

    test_db.add(
        RotationEra(
            start_date=datetime.date(2026, 1, 2),
            end_date=None,
            rotation_length=10,
            weeks_pattern=_ROTATION_ERA_PATTERN,
        )
    )
    test_db.commit()

    yield test_client, test_db

    clear_schedule_cache()


def test_mid_month_change_shows_both_persons(month_env):
    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=anna.id,
        new_user_id=bert.id,
        person_id=3,
        new_name="Bert",
        new_username="bert1",
        effective_from=datetime.date(2026, 8, 15),
        created_by=1,
    )

    resp = client.get("/month?year=2026&month=8")

    assert resp.status_code == 200
    assert "Anna" in resp.text
    assert "Bert" in resp.text


def test_departed_person_absent_in_later_month(month_env):
    """A position vacant for the whole displayed month shows no column at all.

    Anna left with no successor; September falls entirely in the resulting
    gap. Fully vacant positions are hidden (no placeholder column).
    """
    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    end_employment(session, anna.id, 3, end_date=datetime.date(2026, 8, 4))

    resp = client.get("/month?year=2026&month=9")

    assert resp.status_code == 200
    assert "Anna" not in resp.text
    assert "Vakant" not in resp.text and "Vacant" not in resp.text


def test_mid_week_change_shows_both_rows(month_env):
    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    # Thursday of ISO week 34, 2026: Anna holds the position until Wednesday,
    # Bert takes over from Thursday within the same week.
    thursday = datetime.date.fromisocalendar(2026, 34, 4)
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=anna.id,
        new_user_id=bert.id,
        person_id=3,
        new_name="Bert",
        new_username="bert1",
        effective_from=thursday,
        created_by=1,
    )

    resp = client.get("/week?year=2026&week=34")

    assert resp.status_code == 200
    assert "Anna" in resp.text
    assert "Bert" in resp.text


def test_departed_person_absent_in_later_week(month_env):
    """A position vacant for the whole displayed week shows no row at all.

    Anna left with no successor; week 34 falls entirely in the resulting gap.
    Fully vacant positions are hidden (no placeholder row), unlike a partial
    gap within an otherwise-active week, which still needs the OFF cells.
    """
    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    # End employment before ISO week 34, 2026: the position row is fully vacant.
    end_employment(session, anna.id, 3, end_date=datetime.date(2026, 8, 4))

    resp = client.get("/week?year=2026&week=34")

    assert resp.status_code == 200
    assert "Anna" not in resp.text
    assert "Vakant" not in resp.text and "Vacant" not in resp.text


def _year_header_ths(html: str, person_id: int) -> list[str]:
    """Return the full person-header <th> tags for a position's holder columns.

    The year view renders one column per holder segment, keyed by a col_key of
    the form "<person_id>-<user_id>" (or "<person_id>-vacant"), so a position
    can have several header cells. Each returned string is the whole <th ...>
    ... </th> tag including its attributes.
    """
    return re.findall(
        rf'(<th class="person-header[^"]*" data-person="{person_id}-[^"]*".*?</th>)',
        html,
        re.DOTALL,
    )


def _headers_containing(html: str, name: str) -> int:
    """Count <th class="person-header...">...</th> blocks whose content includes name.

    Robust against whitespace between Jinja tags and against the name appearing
    more than once elsewhere on the page (e.g. toggle checkboxes): only the
    single header cell rendered once per row/column is counted, matching the
    year view's col_key format ("<pid>-<uid>" or the merged "user-<uid>")
    without needing to know which format applies.
    """
    blocks = re.findall(r'(<th class="person-header[^"]*"[^>]*>.*?</th>)', html, re.DOTALL)
    return sum(1 for b in blocks if name in b)


def _rows_containing(html: str, row_class: str, name: str) -> int:
    """Count <tr class="{row_class}">...</tr> blocks whose content includes name."""
    blocks = re.findall(rf'(<tr class="{row_class}">.*?</tr>)', html, re.DOTALL)
    return sum(1 for b in blocks if name in b)


def test_year_header_vacant_after_departure(month_env):
    """A position vacant for the whole displayed year shows no column at all.

    Anna held position 3 until 2026-08-04 with no successor. The 2027 view has
    no holder overlapping that year at all, so the position-3 column is fully
    hidden (Goal 2: no vacant placeholder), matching the week and month views.
    """
    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    end_employment(session, anna.id, 3, end_date=datetime.date(2026, 8, 4))

    resp = client.get("/year?year=2027")

    assert resp.status_code == 200
    ths = _year_header_ths(resp.text, 3)
    assert len(ths) == 0
    assert "Anna" not in resp.text
    assert "Vakant" not in resp.text and "Vacant" not in resp.text


def test_year_splits_columns_per_holder_past_hidden(month_env):
    """A mid-year change yields one header column per holder, past one hidden.

    Isak held position 3 until a date in the past relative to today, Omar took
    over the day after. The year view for the current year must render two
    separate position-3 header cells (not a joined one), each linking to its
    holder's own personal year view. Isak's column carries the past marker and
    starts hidden; Omar's does not.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    isak = _make_user(session, 11, "isak1", "Isak")
    omar = _make_user(session, 12, "omar1", "Omar")

    today = get_today()
    isak_end = today - datetime.timedelta(days=10)
    omar_start = isak_end + datetime.timedelta(days=1)

    start_employment(session, isak.id, 3, "Isak", "isak1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=isak.id,
        new_user_id=omar.id,
        person_id=3,
        new_name="Omar",
        new_username="omar1",
        effective_from=omar_start,
        created_by=1,
    )

    # Authenticate as admin so the day drill-down links and totals row render.
    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/year?year={today.year}")

    assert resp.status_code == 200
    ths = _year_header_ths(resp.text, 3)
    assert len(ths) == 2

    isak_th = next(th for th in ths if "Isak" in th)
    omar_th = next(th for th in ths if "Omar" in th)
    # Separate cells, not a joined header.
    assert "Omar" not in isak_th
    assert "Isak" not in omar_th
    # Isak departed in the past: past marker set and column hidden by default.
    assert 'data-past="1"' in isak_th
    assert "display:none" in isak_th
    # Omar is current: no past marker, column visible.
    assert 'data-past="0"' in omar_th
    assert "display:none" not in omar_th
    # Each holder links to their own personal year view.
    assert f'/year/11?year={today.year}"' in isak_th
    assert f'/year/12?year={today.year}"' in omar_th
    # Past holders' filter checkbox is disabled server-side so a departed holder
    # cannot be revealed individually while past days are hidden.
    assert re.search(r'<input[^>]*data-past="1"[^>]*\bdisabled\b', resp.text)


def test_year_future_swap_merges_into_one_visible_column(month_env):
    """A future-dated swap still merges into one column per person, visible now.

    Isak holds position 3 and Omar holds position 5; they swap on a date in the
    future. Before this refactor, a future-dated swap rendered as two separate
    per-position columns (the incoming one hidden until effective) because each
    position was processed independently. Now that segments are grouped by
    user_id across positions first, Isak and Omar each get ONE merged column
    spanning both positions (per the swap-merge design, a genuine swap
    participant's column is never flagged past or future as a whole - unlike a
    genuine departure/successor pair, which stays on separate user_ids and is
    unaffected). The merged column is therefore visible immediately, not hidden
    pending the future swap date.

    Uses fixed dates plus the year view's `?simulated_date=` testing hook
    (rather than monkeypatching get_today) to avoid flakiness: real "today"
    would occasionally land on a probe date where positions 3 and 5 happen to
    share a rotation code, silently weakening or spuriously failing the
    differentiation check below.
    """
    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    isak = _make_user(session, 11, "isak1", "Isak")
    omar = _make_user(session, 12, "omar1", "Omar")

    # Fixed dates, verified to give positions 3 and 5 different rotation codes
    # on both probe days: 2027-06-10 (N1 vs OFF) and 2027-06-20 (N2 vs OC).
    simulated_today = datetime.date(2027, 5, 15)
    swap_date = datetime.date(2027, 6, 15)
    pre_swap_day = datetime.date(2027, 6, 10)
    post_swap_day = datetime.date(2027, 6, 20)

    start_employment(session, isak.id, 3, "Isak", "isak1", datetime.date(2026, 1, 2), created_by=1)
    start_employment(session, omar.id, 5, "Omar", "omar1", datetime.date(2026, 1, 2), created_by=1)
    swap_positions(session, 3, 5, swap_date, created_by=1)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/year?year={simulated_today.year}&simulated_date={simulated_today.isoformat()}")
    assert resp.status_code == 200

    # One column per person, not one per position segment.
    assert _headers_containing(resp.text, "Isak") == 1
    assert _headers_containing(resp.text, "Omar") == 1
    # Neither position number keys a separate header anymore: both people's
    # segments were grouped under their own user_id across positions 3 and 5.
    assert len(_year_header_ths(resp.text, 3)) == 0
    assert len(_year_header_ths(resp.text, 5)) == 0

    blocks = re.findall(r'(<th class="person-header[^"]*"[^>]*>.*?</th>)', resp.text, re.DOTALL)
    isak_th = next(b for b in blocks if "Isak" in b)
    omar_th = next(b for b in blocks if "Omar" in b)
    # Merged swap columns are visible immediately, not force-hidden pending the
    # future swap date.
    assert 'data-past="0"' in isak_th and 'data-future="0"' in isak_th
    assert "display:none" not in isak_th
    assert 'data-past="0"' in omar_th and 'data-future="0"' in omar_th
    assert "display:none" not in omar_th

    # Verify per-day content: check that each person's cell shows their own real
    # shift from whichever position they held on that specific date.
    isak_pre_shift, _ = determine_shift_for_date(pre_swap_day, start_week=3)
    isak_post_shift, _ = determine_shift_for_date(post_swap_day, start_week=5)
    omar_pre_shift, _ = determine_shift_for_date(pre_swap_day, start_week=5)
    omar_post_shift, _ = determine_shift_for_date(post_swap_day, start_week=3)

    for label, link_id, day, expected in [
        ("Isak pre-swap", isak.id, pre_swap_day, isak_pre_shift),
        ("Isak post-swap", isak.id, post_swap_day, isak_post_shift),
        ("Omar pre-swap", omar.id, pre_swap_day, omar_pre_shift),
        ("Omar post-swap", omar.id, post_swap_day, omar_post_shift),
    ]:
        match = re.search(rf'/day/{link_id}/{day.year}/{day.month}/{day.day}".*?</td>', resp.text, re.DOTALL)
        assert match, f"expected a calendar cell for {label} ({day.isoformat()})"
        cell_html = match.group(0)
        expected_code = expected.code if expected else "OFF"
        assert re.search(rf">\s*{re.escape(expected_code)}\s*<", cell_html), f"{label}: {cell_html}"

    # Sanity check: verify that positions 3 and 5 have different rotations on the
    # test days so this assertion is meaningful (not passing by coincidence).
    codes_pos3 = [determine_shift_for_date(d, start_week=3)[0] for d in [pre_swap_day, post_swap_day]]
    codes_pos5 = [determine_shift_for_date(d, start_week=5)[0] for d in [pre_swap_day, post_swap_day]]
    codes_pos3 = [c.code if c else "OFF" for c in codes_pos3]
    codes_pos5 = [c.code if c else "OFF" for c in codes_pos5]
    assert codes_pos3 != codes_pos5, "positions 3 and 5 must differ on the test days for a meaningful check"


def test_year_ongoing_holder_visible_in_later_year(month_env):
    """An ongoing holder viewed in a later year is not mistaken for a future hire.

    Nils holds position 3 with an open record since 2026. Viewing 2027 clamps his
    segment start to 2027-01-01, which is after today, but his raw employment
    start is in the past, so his column must stay visible (future flag off).
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    nils = _make_user(session, 11, "nils1", "Nils")
    start_employment(session, nils.id, 3, "Nils", "nils1", datetime.date(2026, 1, 2), created_by=1)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get("/year?year=2027")
    assert resp.status_code == 200

    ths3 = _year_header_ths(resp.text, 3)
    assert len(ths3) == 1
    assert "Nils" in ths3[0]
    assert 'data-future="0"' in ths3[0]
    assert "display:none" not in ths3[0]


def test_year_future_succession_column_hidden_until_effective(month_env):
    """A future-dated succession (not a swap) still hides the incoming holder.

    Isak currently holds position 3; a future-dated add_person_change hands the
    position to Bob (a different user_id, not a mutual swap). Unlike the
    merged-swap case above, a plain succession keeps its holders on separate
    user_ids, so Bob's column must stay hidden (data-future="1", display:none
    and a disabled filter checkbox) until his start date passes, while Isak's
    current column stays visible.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    isak = _make_user(session, 11, "isak1", "Isak")
    bob = _make_user(session, 12, "bob1", "Bob")

    today = get_today()
    succession_date = today + datetime.timedelta(days=30)

    start_employment(session, isak.id, 3, "Isak", "isak1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=isak.id,
        new_user_id=bob.id,
        person_id=3,
        new_name="Bob",
        new_username="bob1",
        effective_from=succession_date,
        created_by=1,
    )

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/year?year={today.year}")
    assert resp.status_code == 200

    ths = _year_header_ths(resp.text, 3)
    assert len(ths) == 2
    isak_th = next(th for th in ths if "Isak" in th)
    bob_th = next(th for th in ths if "Bob" in th)

    assert 'data-future="0"' in isak_th and "display:none" not in isak_th
    assert 'data-future="1"' in bob_th and "display:none" in bob_th
    assert re.search(r'<input[^>]*data-future="1"[^>]*\bdisabled\b', resp.text)


def _out_of_tenure_cells(html: str, col_key: str) -> list[str]:
    """Return the out-of-tenure day cells (whole <td>...</td>) for a column."""
    return re.findall(
        rf'(<td class="day-cell month-shift-cell out-of-tenure" data-person="{re.escape(col_key)}".*?</td>)',
        html,
        re.DOTALL,
    )


def test_year_out_of_tenure_cells_render_off(month_env):
    """Out-of-tenure day cells render a muted OFF badge instead of being empty."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    isak = _make_user(session, 11, "isak1", "Isak")
    omar = _make_user(session, 12, "omar1", "Omar")

    today = get_today()
    isak_end = today - datetime.timedelta(days=10)
    omar_start = isak_end + datetime.timedelta(days=1)

    start_employment(session, isak.id, 3, "Isak", "isak1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=isak.id,
        new_user_id=omar.id,
        person_id=3,
        new_name="Omar",
        new_username="omar1",
        effective_from=omar_start,
        created_by=1,
    )

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/year?year={today.year}")
    assert resp.status_code == 200

    # Isak's column has out-of-tenure cells for days after his departure; they
    # must show OFF rather than be empty.
    isak_cells = _out_of_tenure_cells(resp.text, "3-11")
    assert isak_cells, "expected out-of-tenure cells for the departed holder"
    assert all("OFF" in cell for cell in isak_cells)
    # The out-of-tenure OFF cell renders the same badge markup and data-act
    # attributes as an in-tenure OFF day cell, so the two are indistinguishable
    # (including under the rotation-mode toggle).
    assert all('class="badge badge-off js-sb" data-act-label="OFF"' in cell for cell in isak_cells)
    assert all('data-orig-code="OFF"' in cell for cell in isak_cells)


def test_year_summary_filters_to_viewed_users_employment(month_env):
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna = _make_user(session, 11, "anna1", "Anna")
    # Peter holds rotation position 3 and is the wage/employment user under test.
    peter = _make_user(session, 12, "peter1", "Peter", person_id=3)
    # Anna held position 3 from rotation start until Peter took over 2026-04-01
    # (add_person_change closes Anna 2026-03-31 and opens Peter 2026-04-01).
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=anna.id,
        new_user_id=peter.id,
        person_id=3,
        new_name="Peter",
        new_username="peter1",
        effective_from=datetime.date(2026, 4, 1),
        created_by=1,
    )

    data = summarize_year_for_person(
        2026,
        3,
        session=session,
        current_user=admin,
        wage_user_id=peter.id,
        employment_user_id=peter.id,
    )
    work_months = [(m["year"], m["month"]) for m in data["months"]]
    assert (2026, 4) in work_months
    # Nothing before Peter's employment start (2026-04) survives the filter,
    # even though the viewer is an admin.
    assert all(not (y < 2026 or (y == 2026 and mo < 4)) for y, mo in work_months)

    # HTTP level: /year/12 as an admin must not render the pre-employment work
    # months (Jan-Mar) while the first employed month (April) is present.
    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/year/12?year=2026")

    assert resp.status_code == 200
    # Closing quote anchors the match so month=1 does not collide with month=10.
    assert '/month/12?year=2026&month=4"' in resp.text
    assert '/month/12?year=2026&month=1"' not in resp.text
    assert '/month/12?year=2026&month=3"' not in resp.text


def test_year_by_user_id_shows_old_holder(month_env):
    """/year/<user_id> for an id <= 10 resolves to that USER, not the position.

    Robin (user id 10) held rotation position 10 until 2026-03-31; Peter (user
    id 12) took over from 2026-04-01. /year/10 must render Robin's data filtered
    to his employment (Jan-Mar), never the successor's later months.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    robin = _make_user(session, 10, "robin1", "Robin")
    peter = _make_user(session, 12, "peter1", "Peter", person_id=10)
    start_employment(session, robin.id, 10, "Robin", "robin1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=robin.id,
        new_user_id=peter.id,
        person_id=10,
        new_name="Peter",
        new_username="peter1",
        effective_from=datetime.date(2026, 4, 1),
        created_by=1,
    )

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/year/10?year=2026")

    assert resp.status_code == 200
    # Robin (user 10) is the subject of the page, not Peter the successor.
    assert "Robin" in resp.text
    assert "Peter" not in resp.text
    # Robin's employment ended 2026-03-31: March present, April filtered out.
    assert '/month/10?year=2026&month=3"' in resp.text
    assert '/month/10?year=2026&month=4"' not in resp.text


def test_year_redirects_non_owner_non_admin(month_env):
    """/year/<id> for someone else's data redirects a regular user to their own page."""
    client, session = month_env
    robin = _make_user(session, 10, "robin1", "Robin")
    start_employment(session, robin.id, 10, "Robin", "robin1", datetime.date(2026, 1, 2), created_by=1)
    viewer = _make_user(session, 13, "viewer1", "Viewer")
    start_employment(session, viewer.id, 3, "Viewer", "viewer1", datetime.date(2026, 1, 2), created_by=1)

    token = create_access_token(data={"sub": str(viewer.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/year/10?year=2026", follow_redirects=False)

    assert resp.status_code == 302
    assert resp.headers["location"] == "/year/13?year=2026"


def test_team_month_links_use_holder_user_ids(month_env):
    """Team month column headers link the holder's user id, not the position.

    A mid-month change at position 3 (Anna user 11 -> Bert user 12) yields two
    columns, each linking its holder's user id. The bare position link /month/3
    must not appear.
    """
    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    add_person_change(
        session,
        old_user_id=anna.id,
        new_user_id=bert.id,
        person_id=3,
        new_name="Bert",
        new_username="bert1",
        effective_from=datetime.date(2026, 8, 15),
        created_by=1,
    )

    resp = client.get("/month?year=2026&month=8")

    assert resp.status_code == 200
    assert "/month/11?year=2026&month=8" in resp.text
    assert "/month/12?year=2026&month=8" in resp.text
    # The changed position is never linked by its bare rotation position.
    assert "/month/3?year=2026&month=8" not in resp.text


def test_team_month_vacant_column_has_no_link(month_env):
    """A vacant position column renders no personal month link."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna = _make_user(session, 11, "anna1", "Anna")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    end_employment(session, anna.id, 3, end_date=datetime.date(2026, 8, 4))

    # Authenticate as admin so day drill-down links render; otherwise the
    # /day/None assertion below is vacuous (day links only render when logged in).
    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get("/month?year=2026&month=9")

    assert resp.status_code == 200
    # Position 3 is vacant in September: neither the departed user nor the bare
    # position is linked for that column.
    assert "/month/11?year=2026&month=9" not in resp.text
    assert "/month/3?year=2026&month=9" not in resp.text
    # Vacant day cells must not emit a broken /day/None drill-down link.
    assert "/day/None" not in resp.text


def test_month_view_merges_swap_into_one_column(month_env):
    """A position swap landing mid-month yields ONE column per person.

    Anna (user 11) and Bert (user 12) hold positions 3 and 5 respectively,
    then swap on 2026-06-15 (mid-June). June's view must show each of them
    exactly once, with their correct shift on each side of the swap date.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, bert.id, 5, "Bert", "bert1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_positions(session, 3, 5, datetime.date(2026, 6, 15), created_by=admin.id)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/month?year=2026&month=6")

    assert resp.status_code == 200
    assert _headers_containing(resp.text, "Anna") == 1
    assert _headers_containing(resp.text, "Bert") == 1


def _fake_month_summary(*, brutto_pay, base_salary, wage_type=WageType.MONTHLY, year=2026, tax_table=None):
    """Minimal fabricated summarize_month_for_person-shaped dict for merge unit tests.

    Only the keys _merge_month_summaries actually reads are populated; every
    numeric/dict field defaults to a zero-ish value via .get() in the merge
    function except "days", which is read unconditionally.
    """
    return {
        "year": year,
        "days": [],
        "brutto_pay": brutto_pay,
        "netto_pay": brutto_pay,
        "base_salary": base_salary,
        "wage_type": wage_type,
        "tax_table": tax_table,
        "absence_details": [],
    }


def test_merge_month_summaries_monthly_wage_does_not_double_count_base(month_env):
    """_merge_month_summaries must not sum a monthly-wage user's flat base twice.

    Two segments share the same 30000 SEK monthly base: segment 1 carries 2000
    SEK of variable pay on top (brutto_pay=32000), segment 2 carries 1500 SEK
    (brutto_pay=31500). The correct merged gross is ONE base plus BOTH
    variable parts (30000 + 2000 + 1500 = 33500). Naively summing the two
    segments' brutto_pay (the pre-fix bug) would instead yield 63500, roughly
    2x the correct value.
    """
    from app.core.schedule.summary import _calculate_tax
    from app.routes.schedule_all import _merge_month_summaries

    base_salary = 30000.0
    seg1 = _fake_month_summary(brutto_pay=base_salary + 2000.0, base_salary=base_salary)
    seg2 = _fake_month_summary(brutto_pay=base_salary + 1500.0, base_salary=base_salary)

    merged = _merge_month_summaries([seg1, seg2])

    naive_double_counted = seg1["brutto_pay"] + seg2["brutto_pay"]
    assert merged["brutto_pay"] == pytest.approx(33500.0)
    assert merged["brutto_pay"] < naive_double_counted - base_salary * 0.9

    expected_netto = merged["brutto_pay"] - _calculate_tax(merged["brutto_pay"], None, payment_year=2026)
    assert merged["netto_pay"] == pytest.approx(expected_netto)


def test_merge_month_summaries_hourly_wage_sums_normally(month_env):
    """An hourly-wage user's already day-derived brutto_pay is safe to sum as-is.

    Hourly brutto_pay carries no flat base component (summary.py replaces it
    entirely with a worked-hours-derived figure), so unlike the monthly case,
    summing two segments' brutto_pay directly is correct.
    """
    from app.routes.schedule_all import _merge_month_summaries

    seg1 = _fake_month_summary(brutto_pay=12000.0, base_salary=30000.0, wage_type=WageType.HOURLY)
    seg2 = _fake_month_summary(brutto_pay=9000.0, base_salary=30000.0, wage_type=WageType.HOURLY)

    merged = _merge_month_summaries([seg1, seg2])

    assert merged["brutto_pay"] == pytest.approx(21000.0)


def test_month_swap_merge_brutto_pay_not_double_counted(month_env):
    """End-to-end: a real monthly-wage swap's merged brutto_pay stays sane.

    Anna holds position 3 from January, swaps to position 5 mid-June. Her June
    column is built from two per-segment summaries (masked to before/after the
    swap) merged by _merge_month_summaries, exercising the exact code path
    show_month_all uses. Before the fix this came out near 2x her monthly wage
    because each segment independently resolved the same flat base salary.
    """
    from app.core.schedule.person_history import get_position_holder_segments
    from app.routes.schedule_all import _merge_month_summaries

    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    start_employment(session, bert.id, 5, "Bert", "bert1", datetime.date(2026, 1, 2), created_by=1)
    swap_positions(session, 3, 5, datetime.date(2026, 6, 15), created_by=1)

    year, month = 2026, 6
    month_start = datetime.date(year, month, 1)
    month_end = datetime.date(year, month, 30)

    segments_by_user: dict[int, list[dict]] = {}
    for pid in (3, 5):
        for seg in get_position_holder_segments(session, pid, month_start, month_end):
            segments_by_user.setdefault(seg["user_id"], []).append({**seg, "person_id": pid})

    anna_segs = sorted(segments_by_user[anna.id], key=lambda s: s["from_date"])
    assert len(anna_segs) == 2, "expected Anna to hold two segments (pre- and post-swap position) in June"

    per_segment_summaries = []
    for seg in anna_segs:
        days = generate_month_data(year, month, seg["person_id"], session=session)
        masked_days = mask_days_to_employment(days, seg["from_date"], seg["to_date"])
        s = summarize_month_for_person(
            year,
            month,
            seg["person_id"],
            session=session,
            year_days=masked_days,
            fetch_tax_table=False,
            payment_year=year,
            wage_user_id=anna.id,
        )
        per_segment_summaries.append(s)

    merged = _merge_month_summaries(per_segment_summaries)

    naive_double_counted = sum(s["brutto_pay"] for s in per_segment_summaries)
    # Well under 2x the base wage: the pre-fix bug summed the flat base once
    # per segment, landing close to naive_double_counted (~2x anna.wage).
    assert merged["brutto_pay"] < naive_double_counted - anna.wage * 0.9
    assert merged["brutto_pay"] < 1.5 * anna.wage


def test_merge_month_summaries_does_not_double_count_whole_month_absence_fields(month_env):
    """_merge_month_summaries must not sum absence-derived fields across segments.

    get_absence_deductions_for_month (wages.py) queries Absence rows by user_id for
    the ENTIRE calendar month with no date-range/segment scoping, so every segment's
    summary independently carries the SAME whole-month sick_days/absence_deduction/
    sick_ob_pay_by_code figures. Naively summing them across two segments (the
    pre-fix bug) doubles them; they must instead be taken from exactly one segment.
    """
    from app.routes.schedule_all import _merge_month_summaries

    seg1 = _fake_month_summary(brutto_pay=30000.0, base_salary=30000.0)
    seg2 = _fake_month_summary(brutto_pay=30000.0, base_salary=30000.0)
    for s in (seg1, seg2):
        s["sick_days"] = 2
        s["sick_hours"] = 16.0
        s["absence_deduction"] = 2400.0
        s["absence_hours"] = 16.0
        s["sick_ob_pay"] = 150.0
        s["sick_total_ob"] = 200.0
        s["sick_ob_lost"] = 50.0
        s["sick_ob_pay_by_code"] = {"kväll": 150.0}
        s["sick_ob_hours_by_code"] = {"kväll": 3.0}

    merged = _merge_month_summaries([seg1, seg2])

    naive_double_counted = seg1["sick_days"] + seg2["sick_days"]
    assert naive_double_counted == 4, "sanity check on the fabricated scenario"
    assert merged["sick_days"] == 2
    assert merged["sick_hours"] == pytest.approx(16.0)
    assert merged["absence_deduction"] == pytest.approx(2400.0)
    assert merged["absence_hours"] == pytest.approx(16.0)
    assert merged["sick_ob_pay"] == pytest.approx(150.0)
    assert merged["sick_total_ob"] == pytest.approx(200.0)
    assert merged["sick_ob_lost"] == pytest.approx(50.0)
    assert merged["sick_ob_pay_by_code"] == {"kväll": 150.0}
    assert merged["sick_ob_hours_by_code"] == {"kväll": 3.0}


def test_merge_month_summaries_parental_days_mix_is_reconstructed_correctly(month_env):
    """parental_days mixes a whole-month absence component with a per-segment
    day-derived (week-based) component; the merge must not double the absence
    part while still summing the week-based part from every segment.

    Fabricated scenario: an identical whole-month absence-derived component of 2
    (constant across both segments, as get_absence_deductions_for_month would
    produce) plus 1 week-based flagged day in segment 1's own `days` list and 2
    in segment 2's. Correct merged total: 2 (absence, once) + 1 + 2 (both
    segments' week-based days) = 5.
    """
    from app.routes.schedule_all import _merge_month_summaries

    seg1 = _fake_month_summary(brutto_pay=30000.0, base_salary=30000.0)
    seg2 = _fake_month_summary(brutto_pay=30000.0, base_salary=30000.0)
    seg1["days"] = [{"parental_leave": True}, {}, {}]
    seg2["days"] = [{}, {"parental_leave": True}, {"parental_leave": True}]
    seg1["parental_days"] = 2 + 1  # absence-only component (2) + this segment's own week day (1)
    seg2["parental_days"] = 2 + 2  # same absence-only component (2) + this segment's 2 week days

    merged = _merge_month_summaries([seg1, seg2])

    assert merged["parental_days"] == 5


def test_month_swap_merge_sick_days_not_double_counted(month_env):
    """End-to-end: a real swap month with sick days on BOTH sides is not double-counted.

    Anna holds position 3 from January, swaps to position 5 mid-June. She has one
    sick day before the swap (still on position 3) and one sick day after the swap
    (now on position 5), both stored as Absence rows keyed by her user_id. Because
    get_absence_deductions_for_month queries the whole calendar month by user_id
    with no segment scoping, each of Anna's two per-segment summaries independently
    reports sick_days=2 for June. Before the fix, _merge_month_summaries summed
    that across both segments to 4; the correct merged total is 2, matching a
    single authoritative call to get_absence_deductions_for_month for the month.
    """
    from app.core.schedule.person_history import get_position_holder_segments
    from app.core.schedule.wages import get_absence_deductions_for_month
    from app.routes.schedule_all import _merge_month_summaries

    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    start_employment(session, bert.id, 5, "Bert", "bert1", datetime.date(2026, 1, 2), created_by=1)
    swap_positions(session, 3, 5, datetime.date(2026, 6, 15), created_by=1)

    # One sick day before the swap (position 3) and one after (position 5).
    session.add(Absence(user_id=anna.id, date=datetime.date(2026, 6, 8), absence_type=AbsenceType.SICK))
    session.add(Absence(user_id=anna.id, date=datetime.date(2026, 6, 22), absence_type=AbsenceType.SICK))
    session.commit()

    year, month = 2026, 6
    month_start = datetime.date(year, month, 1)
    month_end = datetime.date(year, month, 30)

    segments_by_user: dict[int, list[dict]] = {}
    for pid in (3, 5):
        for seg in get_position_holder_segments(session, pid, month_start, month_end):
            segments_by_user.setdefault(seg["user_id"], []).append({**seg, "person_id": pid})

    anna_segs = sorted(segments_by_user[anna.id], key=lambda s: s["from_date"])
    assert len(anna_segs) == 2, "expected Anna to hold two segments (pre- and post-swap position) in June"

    per_segment_summaries = []
    for seg in anna_segs:
        days = generate_month_data(year, month, seg["person_id"], session=session)
        masked_days = mask_days_to_employment(days, seg["from_date"], seg["to_date"])
        s = summarize_month_for_person(
            year,
            month,
            seg["person_id"],
            session=session,
            year_days=masked_days,
            fetch_tax_table=False,
            payment_year=year,
            wage_user_id=anna.id,
        )
        per_segment_summaries.append(s)

    # Sanity check on the bug: each segment independently re-queries the WHOLE
    # month's absences and therefore already reports both sick days.
    assert per_segment_summaries[0]["sick_days"] == 2
    assert per_segment_summaries[1]["sick_days"] == 2

    merged = _merge_month_summaries(per_segment_summaries)

    authoritative = get_absence_deductions_for_month(session, anna.id, year, month, anna.wage)
    assert merged["sick_days"] == authoritative["sick_days"] == 2
    assert merged["absence_deduction"] == pytest.approx(authoritative["total_deduction"])
    naive_double_counted = sum(s["sick_days"] for s in per_segment_summaries)
    assert naive_double_counted == 4
    assert merged["sick_days"] != naive_double_counted


def test_month_swap_merge_brutto_pay_not_double_counted_with_absence(month_env):
    """End-to-end: a real swap month with a paid absence deduction merges brutto_pay once.

    Anna swaps positions mid-June and has one sick day during the month, which
    produces a nonzero absence_deduction from get_absence_deductions_for_month.
    That query is unscoped by segment (whole month, by user_id), so both of
    Anna's per-segment summaries independently carry the SAME absence_deduction
    (and sick_ob_pay). _merge_brutto_netto must fold the flat base_salary AND
    these whole-month absence adjustments in exactly once each, summing only
    each segment's own day-derived variable pay (OB/oncall/OT) on top.
    """
    from app.core.schedule.person_history import get_position_holder_segments
    from app.core.schedule.summary import _calculate_tax
    from app.core.schedule.wages import get_absence_deductions_for_month
    from app.routes.schedule_all import _merge_brutto_netto, _merge_month_summaries

    client, session = month_env
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=1)
    start_employment(session, bert.id, 5, "Bert", "bert1", datetime.date(2026, 1, 2), created_by=1)
    swap_positions(session, 3, 5, datetime.date(2026, 6, 15), created_by=1)

    # One sick day before the swap, still on position 3.
    session.add(Absence(user_id=anna.id, date=datetime.date(2026, 6, 8), absence_type=AbsenceType.SICK))
    session.commit()

    year, month = 2026, 6
    month_start = datetime.date(year, month, 1)
    month_end = datetime.date(year, month, 30)

    segments_by_user: dict[int, list[dict]] = {}
    for pid in (3, 5):
        for seg in get_position_holder_segments(session, pid, month_start, month_end):
            segments_by_user.setdefault(seg["user_id"], []).append({**seg, "person_id": pid})

    anna_segs = sorted(segments_by_user[anna.id], key=lambda s: s["from_date"])
    assert len(anna_segs) == 2, "expected Anna to hold two segments (pre- and post-swap position) in June"

    per_segment_summaries = []
    for seg in anna_segs:
        days = generate_month_data(year, month, seg["person_id"], session=session)
        masked_days = mask_days_to_employment(days, seg["from_date"], seg["to_date"])
        s = summarize_month_for_person(
            year,
            month,
            seg["person_id"],
            session=session,
            year_days=masked_days,
            fetch_tax_table=False,
            payment_year=year,
            wage_user_id=anna.id,
        )
        per_segment_summaries.append(s)

    # Sanity check on the bug: both segments independently carry the SAME
    # whole-month absence deduction (nonzero, since Anna was sick once in June).
    authoritative = get_absence_deductions_for_month(session, anna.id, year, month, anna.wage)
    assert authoritative["total_deduction"] > 0
    assert per_segment_summaries[0]["absence_deduction"] == pytest.approx(authoritative["total_deduction"])
    assert per_segment_summaries[1]["absence_deduction"] == pytest.approx(authoritative["total_deduction"])

    # Independently derive the expected merged gross from the parts summary.py
    # itself defines as genuinely per-segment (OB/oncall/OT pay from each
    # segment's own masked days), plus exactly one base salary and one whole-
    # month absence adjustment - NOT by re-deriving _merge_brutto_netto's formula.
    variable_pay_total = sum(
        sum((s.get("ob_pay") or {}).values()) + (s.get("oncall_pay") or 0) + (s.get("ot_pay") or 0)
        for s in per_segment_summaries
    )
    expected_brutto = (
        anna.wage + variable_pay_total - authoritative["total_deduction"] + authoritative.get("sick_ob_pay", 0.0)
    )

    merged_brutto, merged_netto = _merge_brutto_netto(per_segment_summaries)

    naive_double_counted = per_segment_summaries[0]["brutto_pay"] + (per_segment_summaries[1]["brutto_pay"] - anna.wage)
    assert merged_brutto == pytest.approx(expected_brutto)
    assert merged_brutto != pytest.approx(naive_double_counted)

    expected_netto = merged_brutto - _calculate_tax(merged_brutto, None, payment_year=year)
    assert merged_netto == pytest.approx(expected_netto)

    merged = _merge_month_summaries(per_segment_summaries)
    assert merged["brutto_pay"] == pytest.approx(expected_brutto)


def test_month_view_merge_picks_correct_shift_per_day_across_swap(month_env):
    """The merged month column must show each holder's OWN real shift on each
    side of the swap date, not OFF or the other holder's schedule.

    Anna (position 3) and Bert (position 5) swap on 2026-06-15. Anna's column
    must render position 3's real shift for a pre-swap day and position 5's
    real shift for a post-swap day (and vice versa for Bert).
    """
    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, bert.id, 5, "Bert", "bert1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_positions(session, 3, 5, datetime.date(2026, 6, 15), created_by=admin.id)

    pre_swap_day = datetime.date(2026, 6, 10)
    post_swap_day = datetime.date(2026, 6, 20)
    anna_pre_shift, _ = determine_shift_for_date(pre_swap_day, start_week=3)
    anna_post_shift, _ = determine_shift_for_date(post_swap_day, start_week=5)
    bert_pre_shift, _ = determine_shift_for_date(pre_swap_day, start_week=5)
    bert_post_shift, _ = determine_shift_for_date(post_swap_day, start_week=3)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/month?year=2026&month=6")

    assert resp.status_code == 200

    for label, link_id, day, expected in [
        ("Anna pre-swap", anna.id, pre_swap_day, anna_pre_shift),
        ("Anna post-swap", anna.id, post_swap_day, anna_post_shift),
        ("Bert pre-swap", bert.id, pre_swap_day, bert_pre_shift),
        ("Bert post-swap", bert.id, post_swap_day, bert_post_shift),
    ]:
        match = re.search(rf'/day/{link_id}/{day.year}/{day.month}/{day.day}".*?</td>', resp.text, re.DOTALL)
        assert match, f"expected a calendar cell for {label} ({day.isoformat()})"
        cell_html = match.group(0)
        expected_code = expected.code if expected else "OFF"
        assert re.search(rf">\s*{re.escape(expected_code)}\s*<", cell_html), f"{label}: {cell_html}"


def test_month_view_swap_participant_shows_current_position_badge(month_env):
    """A swap participant's month-view header shows their CURRENT position.

    Anna moves from position 3 to position 5 on 2026-06-15; Bert moves the
    opposite direction. Viewed after the swap has taken effect, each header's
    "(#N)" badge must show the position they hold NOW, matching the year
    view's get_user_person_id-based approach, not the earliest/pre-swap
    position the merge's first segment happens to carry.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, bert.id, 5, "Bert", "bert1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_positions(session, 3, 5, datetime.date(2026, 6, 15), created_by=admin.id)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/month?year=2026&month=6")

    assert resp.status_code == 200

    blocks = re.findall(r'(<th class="person-header"[^>]*>.*?</th>)', resp.text, re.DOTALL)
    anna_th = next(b for b in blocks if "Anna" in b)
    bert_th = next(b for b in blocks if "Bert" in b)

    assert "(#5)" in anna_th and "(#3)" not in anna_th
    assert "(#3)" in bert_th and "(#5)" not in bert_th


def test_month_view_hides_fully_vacant_position(month_env):
    """A position with no holder at all during the displayed month shows no column."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    isak = _make_user(session, 5, "isak1", "Isak")
    start_employment(session, isak.id, 5, "Isak", "isak1", datetime.date(2026, 1, 2), created_by=admin.id)
    end_employment(session, isak.id, 5, datetime.date(2026, 8, 3))

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/month?year=2026&month=8")  # August: Isak leaves Aug 3, rest of month vacant

    assert resp.status_code == 200
    assert "Vakant" not in resp.text and "Vacant" not in resp.text


def test_month_view_orders_columns_by_position_id_with_mixed_legacy(month_env):
    """Column order stays strictly pid-ascending when legacy and history-tracked
    positions are mixed in the same month.

    Position 1 is history-tracked (Alice, via start_employment). Position 5 has
    no PersonHistory at all - a legacy position never touched by the
    person-change admin flow. A two-pass restructure (vacant/legacy columns
    resolved eagerly, per-user merged columns resolved afterwards) must not
    let position 1's column drift after every legacy column: it belongs first.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    alice = _make_user(session, 11, "alice1", "Alice")
    start_employment(session, alice.id, 1, "Alice", "alice1", datetime.date(2026, 1, 2), created_by=admin.id)
    # Position 5 is left completely untouched: zero PersonHistory rows, so it
    # falls into the legacy branch (no segments, no history).

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/month?year=2026&month=1")

    assert resp.status_code == 200
    headers = re.findall(r'(<th class="person-header[^"]*"[^>]*>.*?</th>)', resp.text, re.DOTALL)
    alice_idx = next(i for i, h in enumerate(headers) if "Alice" in h)
    legacy_idx = next(i for i, h in enumerate(headers) if "Person 5" in h)
    assert alice_idx < legacy_idx


def _august_total_ob(year_data: dict) -> float:
    """Return the August (work month 8) total_ob from a personal year summary."""
    for m in year_data["months"]:
        if m["year"] == 2026 and m["month"] == 8:
            return m["total_ob"]
    raise AssertionError("August work month missing from year summary")


def test_mid_month_change_splits_august_ob_between_holders(month_env):
    """A mid-August change must not credit the full month's OB to both holders.

    Anna holds position 3 until 2026-08-14, Bert from 2026-08-15. Each holder's
    personal year summary must count only their own days: the sum of the two
    August OB totals equals the unsplit position OB, and neither holder alone
    carries the whole month.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna = _make_user(session, 11, "anna1", "Anna", person_id=3)
    bert = _make_user(session, 12, "bert1", "Bert", person_id=3)
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=admin.id)
    add_person_change(
        session,
        old_user_id=anna.id,
        new_user_id=bert.id,
        person_id=3,
        new_name="Bert",
        new_username="bert1",
        effective_from=datetime.date(2026, 8, 15),
        created_by=admin.id,
    )

    # Unsplit position OB for August, computed on unmasked days at the holders'
    # wage (both hold wage 30000, so OB pay is identical for either holder).
    unmasked_days = generate_month_data(2026, 8, 3, session=session)
    unsplit = summarize_month_for_person(
        2026, 8, 3, session=session, year_days=unmasked_days, wage_user_id=anna.id, payment_year=2026
    )
    unsplit_ob = sum(float(unsplit["ob_pay"].get(code, 0.0) or 0.0) for code in ("OB1", "OB2", "OB3", "OB4", "OB5"))
    assert unsplit_ob > 0  # the month must carry OB for the split to be meaningful

    anna_year = summarize_year_for_person(
        2026, 3, session=session, current_user=admin, wage_user_id=anna.id, employment_user_id=anna.id
    )
    bert_year = summarize_year_for_person(
        2026, 3, session=session, current_user=admin, wage_user_id=bert.id, employment_user_id=bert.id
    )
    anna_aug = _august_total_ob(anna_year)
    bert_aug = _august_total_ob(bert_year)

    # The two holders partition August OB: their sum reconstructs the full month,
    # and neither holder alone carries all of it.
    assert anna_aug + bert_aug == pytest.approx(unsplit_ob)
    assert anna_aug < unsplit_ob
    assert bert_aug < unsplit_ob


def _seed_future_position_move(session, admin_id):
    """Seed Rickard (user 11) holding position 3 until 2026-09-30, position 8 after.

    The move is future-dated relative to the real today, so date-unaware
    resolution (default today) lands on position 3 while any view of an
    October 2026 date must resolve position 8.
    """
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin_id)
    end_employment(session, rickard.id, 3, end_date=datetime.date(2026, 9, 30))
    start_employment(session, rickard.id, 8, "Rickard", "rickard1", datetime.date(2026, 10, 1), created_by=admin_id)
    return rickard


def test_day_view_resolves_position_by_viewed_date(month_env):
    """/day/<user>/... after a future position move renders the NEW position.

    Rickard moves from position 3 to position 8 on 2026-10-01. Viewing
    2026-10-05 must show position 8's rotation week and shift, not the
    position he holds today.
    """
    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    _seed_future_position_move(session, admin.id)

    viewed = datetime.date(2026, 10, 5)
    shift_new, week_new = determine_shift_for_date(viewed, start_week=8)
    shift_old, week_old = determine_shift_for_date(viewed, start_week=3)
    # Sanity: the two positions must be distinguishable on the chosen date.
    assert week_new != week_old
    assert (shift_new.code if shift_new else "OFF") != (shift_old.code if shift_old else "OFF")

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get("/day/11/2026/10/5")

    assert resp.status_code == 200
    # The rotation-week cell renders exactly "<week>/<length>" for the position
    # held on the viewed date.
    assert f">{week_new}/10<" in resp.text
    assert f">{week_old}/10<" not in resp.text


def test_month_view_shows_real_shifts_from_mid_month_hire_despite_later_move(month_env):
    """A mid-month hire followed by a later position move must not blank the
    hire month.

    Rickard is hired at position 3 on 2026-01-26, then moves to position 8 on
    2026-10-01. Viewing his OWN month page for January 2026 (before either
    record's start date is irrelevant here - the route resolves the position
    using on_date=2026-01-01, which precedes BOTH his records) must still
    render his real position-3 shifts from Jan 26 onward, not OFF for the
    whole month. This reproduces a bug where the fallback in
    get_user_person_id picked the record with the latest effective_from
    (position 8, Oct 1) instead of the earliest one (position 3, Jan 26),
    which made get_employment_period return Oct 1 as the employment start and
    masked the entire month as before-employment.
    """
    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 26), created_by=admin.id)
    end_employment(session, rickard.id, 3, end_date=datetime.date(2026, 9, 30))
    start_employment(session, rickard.id, 8, "Rickard", "rickard1", datetime.date(2026, 10, 1), created_by=admin.id)

    probe = datetime.date(2026, 1, 27)
    real_shift, _ = determine_shift_for_date(probe, start_week=3)
    assert real_shift is not None and real_shift.code != "OFF"  # sanity: position 3 really works that day

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/month/{rickard.id}?year=2026&month=1")

    assert resp.status_code == 200
    assert "Rickard" in resp.text

    match = re.search(rf'/day/{rickard.id}/2026/1/27".*?</td>', resp.text, re.DOTALL)
    assert match, "expected a calendar cell for 2026-01-27"
    cell_html = match.group(0)
    assert re.search(rf">\s*{re.escape(real_shift.code)}\s*<", cell_html)
    assert re.search(r">\s*OFF\s*<", cell_html) is None


def test_month_view_redirects_even_when_successor_exists(month_env):
    """Any viewer is redirected once the month is entirely past the departed
    user's OWN tenure, even when a successor has since taken over the
    position (so the position itself is not vacant).

    Anna holds position 3 Jan 2 - Mar 31, 2026; Bert takes over Apr 1
    (open-ended, immediate succession, no vacancy gap). An admin requesting
    Anna's personal month page for November - long after both her departure
    and Bert's takeover - must be redirected to the team month view, not
    shown a masked calendar. The redirect rule depends solely on Anna's own
    tenure end, never on whether a successor exists.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=admin.id)
    add_person_change(
        session,
        old_user_id=anna.id,
        new_user_id=bert.id,
        person_id=3,
        new_name="Bert",
        new_username="bert1",
        effective_from=datetime.date(2026, 4, 1),
        created_by=admin.id,
    )

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/month/{anna.id}?year=2026&month=11", follow_redirects=False)

    assert resp.status_code == 302
    assert resp.headers["location"] == "/month?year=2026&month=11"


def _first_working_day(dates, start_week):
    """Return the first date in dates whose position (start_week) is a real shift."""
    from app.core.schedule import determine_shift_for_date

    for d in dates:
        shift, _ = determine_shift_for_date(d, start_week=start_week)
        if shift is not None and shift.code != "OFF":
            return d, shift.code
    return None, None


def test_month_view_shows_swap_padding_days_from_next_month(month_env):
    """September's grid padding days from October must show the NEW position.

    Rickard holds position 3 until 2026-09-30, position 8 from 2026-10-01. The
    September calendar grid is expanded to full weeks, pulling in the first few
    October days as padding. Those padding days fall in Rickard's own position-8
    tenure and must render his real position-8 shifts, not OFF.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    _seed_future_position_move(session, admin.id)

    # October padding days shown in the September grid (Oct 1-4 complete the last week).
    padding = [datetime.date(2026, 10, d) for d in range(1, 5)]
    probe, real_code = _first_working_day(padding, start_week=8)
    assert probe is not None, "expected at least one working October padding day for position 8"

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get("/month/11?year=2026&month=9")

    assert resp.status_code == 200
    match = re.search(rf'/day/11/2026/{probe.month}/{probe.day}".*?</td>', resp.text, re.DOTALL)
    assert match, f"expected a calendar cell for {probe.isoformat()}"
    cell_html = match.group(0)
    assert re.search(rf">\s*{re.escape(real_code)}\s*<", cell_html), cell_html
    assert re.search(r">\s*OFF\s*<", cell_html) is None


def test_month_view_shows_swap_padding_days_from_prev_month(month_env):
    """October's grid padding days from September must show the OLD position.

    The October calendar grid pulls in the last few September days as padding.
    Those days fall in Rickard's own position-3 tenure (through 2026-09-30) and
    must render his real position-3 shifts, not OFF.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    _seed_future_position_move(session, admin.id)

    # September padding days shown in the October grid (Sep 28-30 lead the first week).
    padding = [datetime.date(2026, 9, d) for d in (28, 29, 30)]
    probe, real_code = _first_working_day(padding, start_week=3)
    assert probe is not None, "expected at least one working September padding day for position 3"

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get("/month/11?year=2026&month=10")

    assert resp.status_code == 200
    match = re.search(rf'/day/11/2026/{probe.month}/{probe.day}".*?</td>', resp.text, re.DOTALL)
    assert match, f"expected a calendar cell for {probe.isoformat()}"
    cell_html = match.group(0)
    assert re.search(rf">\s*{re.escape(real_code)}\s*<", cell_html), cell_html
    assert re.search(r">\s*OFF\s*<", cell_html) is None


def test_month_view_redirects_for_padding_month_even_with_successor(month_env):
    """A month whose only content would have been successor padding also redirects.

    Anna held position 3 until 2026-03-31; Bert took over. November is
    entirely past Anna's own tenure end, so any viewer (here, an admin) is
    redirected to the team month view before the grid - and its December
    padding days, which would otherwise belong to Bert - is ever built. This
    supersedes the old 313a7df successor-leak fix's masking assertion for
    this exact request: since the redirect now fires first, there is no page
    left to leak from.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    _anna, _bert = _seed_departed_with_successor(session, admin.id)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get("/month/11?year=2026&month=11", follow_redirects=False)

    assert resp.status_code == 302
    assert resp.headers["location"] == "/month?year=2026&month=11"


def _seed_departed_with_successor(session, admin_id, *, old_end=datetime.date(2026, 3, 31)):
    """Seed Anna holding position 3 until old_end, Bert taking over the next day.

    Immediate succession, no vacancy gap - Anna's own page for any date after
    old_end must not render Bert's real schedule.
    """
    anna = _make_user(session, 11, "anna1", "Anna")
    bert = _make_user(session, 12, "bert1", "Bert")
    start_employment(session, anna.id, 3, "Anna", "anna1", datetime.date(2026, 1, 2), created_by=admin_id)
    add_person_change(
        session,
        old_user_id=anna.id,
        new_user_id=bert.id,
        person_id=3,
        new_name="Bert",
        new_username="bert1",
        effective_from=old_end + datetime.timedelta(days=1),
        created_by=admin_id,
        old_end_date=old_end,
    )
    return anna, bert


def test_day_view_masks_successor_schedule_after_departure(month_env):
    """A departed user's day page must not leak a successor's real shift."""
    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna, _bert = _seed_departed_with_successor(session, admin.id)

    probe = datetime.date(2026, 11, 16)
    real_shift, _ = determine_shift_for_date(probe, start_week=3)
    assert real_shift is not None and real_shift.code != "OFF"

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/day/{anna.id}/2026/11/16")

    assert resp.status_code == 200
    assert "Anna" in resp.text
    assert re.search(rf">\s*{re.escape(real_shift.code)}\s*<", resp.text) is None


def test_week_view_redirects_even_when_successor_exists(month_env):
    """Any viewer is redirected once the week is entirely past the departed
    user's OWN tenure, even when a successor has since taken over the
    position (so the position itself is not vacant).

    Anna holds position 3 Jan 2 - Mar 31, 2026; Bert takes over Apr 1
    (open-ended, immediate succession, no vacancy gap). An admin requesting
    Anna's personal week page for mid-November - long after both her
    departure and Bert's takeover - must be redirected to the team week
    view, not shown a masked week. The redirect rule depends solely on
    Anna's own tenure end, never on whether a successor exists. This
    supersedes the old test_week_view_masks_successor_schedule_after_departure
    masking assertion for this exact request: since the redirect now fires
    first, there is no page left to leak from.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna, _bert = _seed_departed_with_successor(session, admin.id)

    iso_year, iso_week, _ = datetime.date(2026, 11, 16).isocalendar()

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/week/{anna.id}?year={iso_year}&week={iso_week}", follow_redirects=False)

    assert resp.status_code == 302
    assert resp.headers["location"] == f"/week?year={iso_year}&week={iso_week}"


def test_excel_export_masks_successor_schedule_after_departure(month_env):
    """The month Excel export for a departed user must not contain a successor's hours."""
    import io as _io

    import openpyxl

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    anna, _bert = _seed_departed_with_successor(session, admin.id)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get(f"/month/{anna.id}/export-excel?year=2026&month=11")

    assert resp.status_code == 200
    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    ws = wb.active
    rows = [row for row in ws.iter_rows(values_only=True) if str(row[0]).startswith("2026-11")]
    assert len(rows) == 30, "expected one row per day in November 2026"
    # Every November row must show "Ledig" (OFF) - no successor shift leaking in.
    for row in rows:
        assert row[2] == "Ledig"


def test_excel_export_resolves_position_by_exported_month(month_env):
    """The month Excel export after a future move exports the NEW position.

    Exporting October 2026 for Rickard must contain position 8's shift on a
    day where positions 3 and 8 differ.
    """
    import io as _io

    import openpyxl

    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    _seed_future_position_move(session, admin.id)

    # The sheet's Skifttyp column carries the shift LABEL, so compare labels.
    probe = datetime.date(2026, 10, 5)
    shift_new, _ = determine_shift_for_date(probe, start_week=8)
    shift_old, _ = determine_shift_for_date(probe, start_week=3)
    expected = shift_new.label if shift_new else "OFF"
    wrong = shift_old.label if shift_old else "OFF"
    assert expected != wrong  # sanity: the probe day distinguishes the positions

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    resp = client.get("/month/11/export-excel?year=2026&month=10")

    assert resp.status_code == 200
    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    ws = wb.active
    # Find the probe day's row (column 1 = date, column 3 = shift code).
    probe_rows = [row for row in ws.iter_rows(values_only=True) if str(row[0]).startswith("2026-10-05")]
    assert probe_rows, "expected a row for 2026-10-05 in the export"
    assert probe_rows[0][2] == expected


def _ob_total(summary: dict) -> float:
    """Sum the five OB pay codes of a month summary."""
    ob_pay = summary.get("ob_pay", {}) or {}
    return sum(float(ob_pay.get(code, 0.0) or 0.0) for code in ("OB1", "OB2", "OB3", "OB4", "OB5"))


def test_year_summary_spans_position_swap(month_env):
    """A user's year summary covers BOTH positions of a future-dated move.

    Rickard holds position 3 until 2026-09-30 and position 8 from 2026-10-01.
    His personal year summary must contain September (from position 3) and
    October (from position 8) work months, each matching that position's own
    unsplit month summary.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _seed_future_position_move(session, admin.id)

    year_data = summarize_year_for_person(
        2026, 3, session=session, current_user=admin, wage_user_id=rickard.id, employment_user_id=rickard.id
    )
    by_month = {(m["year"], m["month"]): m for m in year_data["months"]}
    assert (2026, 9) in by_month
    assert (2026, 10) in by_month

    sep_ref = summarize_month_for_person(
        2026,
        9,
        3,
        session=session,
        year_days=generate_month_data(2026, 9, 3, session=session),
        wage_user_id=rickard.id,
        payment_year=2026,
    )
    oct_ref = summarize_month_for_person(
        2026,
        10,
        8,
        session=session,
        year_days=generate_month_data(2026, 10, 8, session=session),
        wage_user_id=rickard.id,
        payment_year=2026,
    )
    assert sep_ref["num_shifts"] > 0
    assert oct_ref["num_shifts"] > 0
    assert by_month[(2026, 9)]["num_shifts"] == sep_ref["num_shifts"]
    assert by_month[(2026, 10)]["num_shifts"] == oct_ref["num_shifts"]
    assert by_month[(2026, 9)]["total_ob"] == pytest.approx(_ob_total(sep_ref))
    assert by_month[(2026, 10)]["total_ob"] == pytest.approx(_ob_total(oct_ref))

    # HTTP level: the personal year page lists months across both positions.
    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/year/11?year=2026")

    assert resp.status_code == 200
    assert '/month/11?year=2026&month=9"' in resp.text
    assert '/month/11?year=2026&month=10"' in resp.text


def test_year_summary_stitches_mid_month_position_move(month_env):
    """A mid-month position move stitches the transition month from both halves.

    Rickard holds position 3 until 2026-08-14 and position 8 from 2026-08-15.
    His August work month must equal position 3's masked first half plus
    position 8's masked second half, and later months come from position 8.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin.id)
    end_employment(session, rickard.id, 3, end_date=datetime.date(2026, 8, 14))
    start_employment(session, rickard.id, 8, "Rickard", "rickard1", datetime.date(2026, 8, 15), created_by=admin.id)

    # Reference halves: each position's August, masked to its tenure segment.
    a_days = mask_days_to_employment(
        generate_month_data(2026, 8, 3, session=session), datetime.date(2026, 8, 1), datetime.date(2026, 8, 14)
    )
    a_sum = summarize_month_for_person(
        2026, 8, 3, session=session, year_days=a_days, wage_user_id=rickard.id, payment_year=2026
    )
    b_days = mask_days_to_employment(
        generate_month_data(2026, 8, 8, session=session), datetime.date(2026, 8, 15), datetime.date(2026, 8, 31)
    )
    b_sum = summarize_month_for_person(
        2026, 8, 8, session=session, year_days=b_days, wage_user_id=rickard.id, payment_year=2026
    )
    # Both halves must carry OB for the reconstruction to be meaningful.
    assert _ob_total(a_sum) > 0
    assert _ob_total(b_sum) > 0

    year_data = summarize_year_for_person(
        2026, 3, session=session, current_user=admin, wage_user_id=rickard.id, employment_user_id=rickard.id
    )
    by_month = {(m["year"], m["month"]): m for m in year_data["months"]}
    aug = by_month[(2026, 8)]

    # The stitched month reconstructs the sum of the two masked halves.
    assert aug["total_ob"] == pytest.approx(_ob_total(a_sum) + _ob_total(b_sum))
    assert aug["num_shifts"] == a_sum["num_shifts"] + b_sum["num_shifts"]
    assert aug["total_hours"] == pytest.approx(a_sum["total_hours"] + b_sum["total_hours"])

    # Months after the move exist and come from position 8.
    assert (2026, 9) in by_month
    sep_ref = summarize_month_for_person(
        2026,
        9,
        8,
        session=session,
        year_days=generate_month_data(2026, 9, 8, session=session),
        wage_user_id=rickard.id,
        payment_year=2026,
    )
    assert sep_ref["num_shifts"] > 0
    assert by_month[(2026, 9)]["num_shifts"] == sep_ref["num_shifts"]


def test_year_summary_counts_user_keyed_absences(month_env):
    """Absence deductions in the user-scoped year summary key on the USER id.

    Rickard is user 11 at rotation position 3. His sick day must appear in his
    own year summary even though his user id differs from the position id
    (absences are stored per user id).
    """
    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin.id)

    # First working day in March 2026 for position 3.
    sick_date = None
    for day in range(1, 32):
        d = datetime.date(2026, 3, day)
        shift, _ = determine_shift_for_date(d, start_week=3)
        if shift and shift.code in ("N1", "N2", "N3"):
            sick_date = d
            break
    assert sick_date is not None

    session.add(Absence(user_id=rickard.id, date=sick_date, absence_type=AbsenceType.SICK))
    session.commit()

    year_data = summarize_year_for_person(
        2026, 3, session=session, current_user=admin, wage_user_id=rickard.id, employment_user_id=rickard.id
    )
    march = next(m for m in year_data["months"] if (m["year"], m["month"]) == (2026, 3))

    assert march["sick_days"] == 1
    assert march["absence_deduction"] > 0


def test_ot_shift_stays_on_pre_swap_holder(month_env):
    """An OT shift dated before a position swap stays on its original holder.

    Rickard (user 11) holds position 3 and Okan (user 8) holds position 8; they
    swap on 2026-10-01. Okan works an overtime shift on 2026-09-10, while he still
    holds position 8. In the September team view that OT must render under position
    8 (Okan), never under position 3 (Rickard's September column).

    Regression: the batch OT fetch keyed rows by the user's CURRENT position (from
    User.person_id, which the swap updated immediately), so Okan's pre-swap OT
    landed on Rickard's column.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    okan = _make_user(session, 8, "okan1", "Okan")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, okan.id, 8, "Okan", "okan1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_positions(session, 3, 8, datetime.date(2026, 10, 1), created_by=admin.id)

    ot_date = datetime.date(2026, 9, 10)
    session.add(
        OvertimeShift(
            user_id=okan.id,
            date=ot_date,
            start_time=datetime.time(18, 0),
            end_time=datetime.time(22, 0),
            hours=4.0,
            ot_pay=1000.0,
            is_extension=False,
        )
    )
    session.commit()

    days = generate_month_data(2026, 9, session=session)
    day = next(d for d in days if d["date"] == ot_date)
    persons = {p["person_id"]: p for p in day["persons"]}

    okan_shift = persons[8]["shift"]
    rickard_shift = persons[3]["shift"]
    # Okan (position 8 in September) carries the OT; Rickard (position 3) does not.
    assert okan_shift is not None and okan_shift.code == "OT"
    assert rickard_shift is None or rickard_shift.code != "OT"


def test_week_view_merges_swap_into_one_row(month_env):
    """A position swap between two active people yields ONE row per person,
    not one row per position segment.

    Rickard (user 11, position 3) and Okan (user 8, position 8) swap on
    2026-10-01. Week 40 2026 (2026-09-28 to 2026-10-04) straddles the swap.
    Each person must appear exactly once, with their own real shifts on each
    side of the swap date - not twice (once per position).
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    okan = _make_user(session, 8, "okan1", "Okan")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, okan.id, 8, "Okan", "okan1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_positions(session, 3, 8, datetime.date(2026, 10, 1), created_by=admin.id)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/week?year=2026&week=40")

    assert resp.status_code == 200
    assert _rows_containing(resp.text, "person-row", "Rickard") == 1
    assert _rows_containing(resp.text, "person-row", "Okan") == 1


def test_week_view_merge_picks_correct_position_per_day_across_swap(month_env):
    """The merged swap row pulls each day's cell from the position actually
    held on that specific date, not from a single fixed position for the
    whole week.

    Same setup as test_week_view_merges_swap_into_one_row: Rickard (position 3)
    and Okan (position 8) swap on 2026-10-01, and week 40 2026 (Mon 2026-09-28
    to Sun 2026-10-04) straddles the boundary. This calls _build_person_rows
    directly (the per-day segment lookup that is the riskiest part of the
    merge refactor) and checks, for every day of the week, that each person's
    cell carries the right person_id AND the right shift code: position 3's
    rotation for Rickard / position 8's for Okan on the days before the swap,
    and the reverse from 2026-10-01 onward.
    """
    from app.core.schedule import determine_shift_for_date
    from app.routes.schedule_all import _build_person_rows

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    okan = _make_user(session, 8, "okan1", "Okan")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, okan.id, 8, "Okan", "okan1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_date = datetime.date(2026, 10, 1)
    swap_positions(session, 3, 8, swap_date, created_by=admin.id)

    monday = datetime.date.fromisocalendar(2026, 40, 1)
    sunday = monday + datetime.timedelta(days=6)
    days_in_week = build_week_data(2026, 40, session=session)
    person_rows = _build_person_rows(session, days_in_week, monday, sunday)

    rickard_row = next(r for r in person_rows if r["person_name"] == "Rickard")
    okan_row = next(r for r in person_rows if r["person_name"] == "Okan")

    for day, cell in zip(days_in_week, rickard_row["cells"], strict=True):
        expected_pid = 3 if day["date"] < swap_date else 8
        expected_shift, _ = determine_shift_for_date(day["date"], start_week=expected_pid)
        expected_code = expected_shift.code if expected_shift else "OFF"
        assert cell is not None, f"Rickard: missing cell for {day['date']}"
        assert cell["person_id"] == expected_pid, f"Rickard: wrong position on {day['date']}"
        actual_code = cell["shift"].code if cell["shift"] else "OFF"
        assert actual_code == expected_code, f"Rickard: {day['date']} expected {expected_code} got {actual_code}"

    for day, cell in zip(days_in_week, okan_row["cells"], strict=True):
        expected_pid = 8 if day["date"] < swap_date else 3
        expected_shift, _ = determine_shift_for_date(day["date"], start_week=expected_pid)
        expected_code = expected_shift.code if expected_shift else "OFF"
        assert cell is not None, f"Okan: missing cell for {day['date']}"
        assert cell["person_id"] == expected_pid, f"Okan: wrong position on {day['date']}"
        actual_code = cell["shift"].code if cell["shift"] else "OFF"
        assert actual_code == expected_code, f"Okan: {day['date']} expected {expected_code} got {actual_code}"

    # Sanity: the two positions must actually differ somewhere in the week for
    # this to be a meaningful check (otherwise a wrong-position bug could pass
    # by coincidence).
    codes_pos3 = [(determine_shift_for_date(d["date"], start_week=3)[0] or None) for d in days_in_week]
    codes_pos8 = [(determine_shift_for_date(d["date"], start_week=8)[0] or None) for d in days_in_week]
    codes_pos3 = [c.code if c else "OFF" for c in codes_pos3]
    codes_pos8 = [c.code if c else "OFF" for c in codes_pos8]
    assert codes_pos3 != codes_pos8


def test_week_view_orders_rows_by_position_id_with_mixed_legacy(month_env):
    """Row order stays strictly pid-ascending when legacy and history-tracked
    positions are mixed in the same week.

    Position 1 is history-tracked (Alice, via start_employment). Position 5 has
    no PersonHistory at all - a legacy position never touched by the
    person-change admin flow - so it falls into the "no row" fast path... no,
    into the legacy branch (no segments, no history) that renders straight from
    the base rotation cells. Before this refactor, a single per-pid loop from 1
    to 10 rendered rows in strict ascending person_id order regardless of
    vacant/legacy/single/succession status. The merge refactor must preserve
    that: position 1's row must render above position 5's, not after every
    legacy row.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    alice = _make_user(session, 11, "alice1", "Alice")
    start_employment(session, alice.id, 1, "Alice", "alice1", datetime.date(2026, 1, 2), created_by=admin.id)
    # Position 5 is left completely untouched: no start_employment/add_person_change
    # calls, so it has zero PersonHistory rows and falls into the legacy branch.

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/week?year=2026&week=3")

    assert resp.status_code == 200
    rows = re.findall(r'(<tr class="person-row">.*?</tr>)', resp.text, re.DOTALL)
    alice_idx = next(i for i, r in enumerate(rows) if "Alice" in r)
    legacy_idx = next(i for i, r in enumerate(rows) if "Person 5" in r)
    assert alice_idx < legacy_idx


def test_week_view_hides_fully_vacant_position(month_env):
    """A position with no holder at all during the displayed week shows no row."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    isak = _make_user(session, 5, "isak1", "Isak")
    start_employment(session, isak.id, 5, "Isak", "isak1", datetime.date(2026, 1, 2), created_by=admin.id)
    end_employment(session, isak.id, 5, end_date=datetime.date(2026, 8, 3))
    # Position 5 has a real gap: Isak left 2026-08-03, nobody holds it until
    # a successor (not seeded here) starts 2026-09-01. Week 35 falls in the gap.

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/week?year=2026&week=35")

    assert resp.status_code == 200
    assert "Vakant" not in resp.text and "Vacant" not in resp.text


def test_year_view_merges_swap_into_one_column(month_env):
    """A position swap between two active people yields ONE column per person
    in the year view, with the correct shift on each side of the swap date."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    okan = _make_user(session, 8, "okan1", "Okan")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, okan.id, 8, "Okan", "okan1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_positions(session, 3, 8, datetime.date(2026, 10, 1), created_by=admin.id)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/year?year=2026")

    assert resp.status_code == 200
    assert _headers_containing(resp.text, "Rickard") == 1
    assert _headers_containing(resp.text, "Okan") == 1


def test_year_view_merge_picks_correct_position_per_day_across_swap(month_env):
    """The merged year column pulls each day's cell from the position actually
    held on that specific date, not from a single fixed position for the
    whole year.

    Rickard (position 3) and Okan (position 8) swap on 2026-10-01. This checks,
    for a day before and a day after the swap, that each person's cell shows
    their own real shift from whichever position they held that day.
    """
    from app.core.schedule import determine_shift_for_date

    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    rickard = _make_user(session, 11, "rickard1", "Rickard")
    okan = _make_user(session, 8, "okan1", "Okan")
    start_employment(session, rickard.id, 3, "Rickard", "rickard1", datetime.date(2026, 1, 2), created_by=admin.id)
    start_employment(session, okan.id, 8, "Okan", "okan1", datetime.date(2026, 1, 2), created_by=admin.id)
    swap_positions(session, 3, 8, datetime.date(2026, 10, 1), created_by=admin.id)

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/year?year=2026")
    assert resp.status_code == 200

    pre_swap_day = datetime.date(2026, 9, 20)
    post_swap_day = datetime.date(2026, 10, 10)
    rickard_pre, _ = determine_shift_for_date(pre_swap_day, start_week=3)
    rickard_post, _ = determine_shift_for_date(post_swap_day, start_week=8)
    okan_pre, _ = determine_shift_for_date(pre_swap_day, start_week=8)
    okan_post, _ = determine_shift_for_date(post_swap_day, start_week=3)

    for label, link_id, day, expected in [
        ("Rickard pre-swap", rickard.id, pre_swap_day, rickard_pre),
        ("Rickard post-swap", rickard.id, post_swap_day, rickard_post),
        ("Okan pre-swap", okan.id, pre_swap_day, okan_pre),
        ("Okan post-swap", okan.id, post_swap_day, okan_post),
    ]:
        match = re.search(rf'/day/{link_id}/{day.year}/{day.month}/{day.day}".*?</td>', resp.text, re.DOTALL)
        assert match, f"expected a calendar cell for {label} ({day.isoformat()})"
        cell_html = match.group(0)
        expected_code = expected.code if expected else "OFF"
        assert re.search(rf">\s*{re.escape(expected_code)}\s*<", cell_html), f"{label}: {cell_html}"


def test_year_view_hides_fully_vacant_position(month_env):
    """A position with no holder at all during the displayed year shows no column."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    isak = _make_user(session, 5, "isak1", "Isak")
    start_employment(session, isak.id, 5, "Isak", "isak1", datetime.date(2026, 1, 2), created_by=admin.id)
    end_employment(session, isak.id, 5, datetime.date(2026, 1, 31))
    # No successor at all after Isak's departure: position 5 has zero overlap
    # with 2027, so its column is fully hidden (Goal 2: no vacant placeholder).

    token = create_access_token(data={"sub": str(admin.id)})
    client.cookies.set("access_token", f"Bearer {token}")
    resp = client.get("/year?year=2027")  # Position 5 has zero overlap with 2027 at all

    assert resp.status_code == 200
    assert "Vakant" not in resp.text and "Vacant" not in resp.text


def test_year_totals_api_rejects_foreign_user_id(month_env):
    """Non-admin totals scoping is limited to legitimate holders of the position.

    A non-admin viewer at position 3 who passes ?user_id pointing at a user who
    never held position 3 gets 403 (otherwise they could back out that user's
    wage level). The request for the real holder still returns totals.
    """
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    # Peter is the non-admin viewer and the legitimate holder of position 3.
    peter = _make_user(session, 12, "peter1", "Peter", person_id=3)
    # Stranger exists but never held position 3.
    _make_user(session, 13, "stranger1", "Stranger", person_id=5)
    start_employment(session, peter.id, 3, "Peter", "peter1", datetime.date(2026, 1, 2), created_by=admin.id)

    token = create_access_token(data={"sub": str(peter.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    # Foreign holder: forbidden.
    forbidden = client.get("/api/year/2026/totals/3?user_id=13")
    assert forbidden.status_code == 403

    # Legitimate holder (Peter himself): totals returned.
    allowed = client.get("/api/year/2026/totals/3?user_id=12")
    assert allowed.status_code == 200
    assert "total_ob" in allowed.json()


def test_month_redirects_departed_user_to_team_view(month_env):
    """A departed user's personal month view redirects to month_all once the
    ENTIRE requested month is after their own last working day."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    robin = _make_user(session, 10, "robin1", "Robin")
    start_employment(session, robin.id, 10, "Robin", "robin1", datetime.date(2026, 1, 2), created_by=admin.id)
    end_employment(session, robin.id, 10, datetime.date(2026, 3, 31))

    token = create_access_token(data={"sub": str(robin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    # July is entirely after his own tenure: redirect.
    resp = client.get("/month/10?year=2026&month=7", follow_redirects=False)
    assert resp.status_code == 302
    assert resp.headers["location"] == "/month?year=2026&month=7"

    # March is his own real last month: no redirect, renders normally.
    resp2 = client.get("/month/10?year=2026&month=3", follow_redirects=False)
    assert resp2.status_code == 200


def test_week_redirects_departed_user_to_team_view(month_env):
    """A departed user's personal week view redirects to week_all once the
    ENTIRE requested week is after their own last working day."""
    client, session = month_env
    admin = _make_user(session, 2, "admin1", "Admin", role=UserRole.ADMIN)
    robin = _make_user(session, 10, "robin1", "Robin")
    start_employment(session, robin.id, 10, "Robin", "robin1", datetime.date(2026, 1, 2), created_by=admin.id)
    end_employment(session, robin.id, 10, datetime.date(2026, 3, 31))

    token = create_access_token(data={"sub": str(robin.id)})
    client.cookies.set("access_token", f"Bearer {token}")

    # Week 30 (late July) is entirely after his tenure: redirect.
    resp = client.get("/week/10?year=2026&week=30", follow_redirects=False)
    assert resp.status_code == 302
    assert resp.headers["location"] == "/week?year=2026&week=30"

    # Week 1 (his own real first week) still renders normally.
    resp2 = client.get("/week/10?year=2026&week=1", follow_redirects=False)
    assert resp2.status_code == 200
